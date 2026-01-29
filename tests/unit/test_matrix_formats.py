"""Tests for jamb.matrix.formats module."""

import csv
import io
import json

import pytest
from openpyxl import load_workbook

from jamb.core.models import (
    ChainRow,
    FullChainMatrix,
    Item,
    LinkedTest,
    MatrixMetadata,
    TestEnvironment,
    TestRecord,
)
from jamb.matrix.formats.csv import render_full_chain_csv, render_test_records_csv
from jamb.matrix.formats.html import render_full_chain_html, render_test_records_html
from jamb.matrix.formats.json import render_full_chain_json, render_test_records_json
from jamb.matrix.formats.markdown import (
    render_full_chain_markdown,
    render_test_records_markdown,
)
from jamb.matrix.formats.xlsx import render_full_chain_xlsx, render_test_records_xlsx


@pytest.fixture
def sample_test_records():
    """Sample test records for testing renderers."""
    return [
        TestRecord(
            test_id="TC001",
            test_name="test_valid",
            test_nodeid="test_auth.py::test_valid",
            outcome="passed",
            requirements=["SRS001", "SRS002"],
            test_actions=["Enter credentials", "Click submit"],
            expected_results=["Login successful"],
            actual_results=["Login successful"],
            notes=["Verified with valid user"],
            execution_timestamp="2024-01-01T10:00:00Z",
        ),
        TestRecord(
            test_id="TC002",
            test_name="test_invalid",
            test_nodeid="test_auth.py::test_invalid",
            outcome="failed",
            requirements=["SRS003"],
            test_actions=["Enter invalid credentials"],
            expected_results=["Error message displayed"],
            actual_results=["No error shown"],
            notes=["[FAILURE] AssertionError: expected error message"],
            execution_timestamp="2024-01-01T10:01:00Z",
        ),
    ]


@pytest.fixture
def sample_metadata():
    """Sample metadata for testing renderers."""
    return MatrixMetadata(
        software_version="1.0.0",
        tester_id="CI Pipeline",
        execution_timestamp="2024-01-01T10:00:00Z",
        environment=TestEnvironment(
            os_name="Linux",
            os_version="5.4.0",
            python_version="3.10.0",
            platform="x86_64",
            processor="Intel",
            hostname="ci-server",
            cpu_count=4,
            test_tools={"pytest": "7.0.0"},
        ),
    )


@pytest.fixture
def sample_full_chain_matrices():
    """Sample full chain matrices for testing renderers."""
    item1 = Item(uid="SYS001", text="System requirement", document_prefix="SYS")
    item2 = Item(
        uid="SRS001",
        text="Software requirement",
        document_prefix="SRS",
        header="Login",
    )

    row1 = ChainRow(
        chain={"SYS": item1, "SRS": item2},
        rollup_status="Passed",
        descendant_tests=[
            LinkedTest(
                test_nodeid="test_auth.py::test_login",
                item_uid="SRS001",
                test_outcome="passed",
            )
        ],
        ancestor_uids=["PRJ001"],
    )
    row2 = ChainRow(
        chain={"SYS": item1, "SRS": None},
        rollup_status="Not Covered",
        descendant_tests=[],
        ancestor_uids=["PRJ001"],
    )

    matrix = FullChainMatrix(
        path_name="SYS -> SRS",
        document_hierarchy=["SYS", "SRS"],
        rows=[row1, row2],
        summary={"total": 2, "passed": 1, "failed": 0, "not_covered": 1},
        include_ancestors=True,
    )
    return [matrix]


# =============================================================================
# Test Records HTML Tests
# =============================================================================


class TestRenderTestRecordsHtml:
    """Tests for test records HTML renderer."""

    def test_renders_basic_structure(self, sample_test_records):
        """Test that HTML output has basic structure."""
        html = render_test_records_html(sample_test_records)

        assert "<html" in html
        assert "<table" in html
        assert "</html>" in html

    def test_contains_charset_meta_tag(self, sample_test_records):
        """Test that HTML output declares UTF-8 charset."""
        html = render_test_records_html(sample_test_records)

        assert '<meta charset="utf-8">' in html

    def test_contains_test_id(self, sample_test_records):
        """Test that HTML contains test ID."""
        html = render_test_records_html(sample_test_records)

        assert "TC001" in html
        assert "TC002" in html

    def test_contains_test_name(self, sample_test_records):
        """Test that HTML contains test name."""
        html = render_test_records_html(sample_test_records)

        assert "test_valid" in html
        assert "test_invalid" in html

    def test_passed_status_class(self, sample_test_records):
        """Test that passed tests have correct CSS class."""
        html = render_test_records_html(sample_test_records)

        assert 'class="passed"' in html

    def test_failed_status_class(self, sample_test_records):
        """Test that failed tests have correct CSS class."""
        html = render_test_records_html(sample_test_records)

        assert 'class="failed"' in html

    def test_includes_metadata(self, sample_test_records, sample_metadata):
        """Test that HTML includes metadata when provided."""
        html = render_test_records_html(sample_test_records, metadata=sample_metadata)

        assert "1.0.0" in html  # software version
        assert "CI Pipeline" in html  # tester id

    def test_includes_requirements(self, sample_test_records):
        """Test that requirements are included."""
        html = render_test_records_html(sample_test_records)

        assert "SRS001" in html
        assert "SRS002" in html

    def test_includes_test_actions(self, sample_test_records):
        """Test that test actions are included."""
        html = render_test_records_html(sample_test_records)

        assert "Enter credentials" in html


# =============================================================================
# Full Chain HTML Tests
# =============================================================================


class TestRenderFullChainHtml:
    """Tests for full chain HTML renderer."""

    def test_renders_basic_structure(self, sample_full_chain_matrices):
        """Test that HTML output has basic structure."""
        html = render_full_chain_html(sample_full_chain_matrices)

        assert "<html" in html
        assert "<table" in html
        assert "</html>" in html

    def test_contains_document_columns(self, sample_full_chain_matrices):
        """Test that HTML contains document hierarchy columns."""
        html = render_full_chain_html(sample_full_chain_matrices)

        assert "SYS" in html
        assert "SRS" in html

    def test_contains_item_uid(self, sample_full_chain_matrices):
        """Test that HTML contains item UIDs."""
        html = render_full_chain_html(sample_full_chain_matrices)

        assert "SYS001" in html
        assert "SRS001" in html

    def test_contains_status(self, sample_full_chain_matrices):
        """Test that HTML contains status."""
        html = render_full_chain_html(sample_full_chain_matrices)

        assert "Passed" in html
        assert "Not Covered" in html

    def test_tc_mapping_applied(self, sample_full_chain_matrices):
        """Test that TC mapping is applied to test display."""
        tc_mapping = {"test_auth.py::test_login": "TC001"}
        html = render_full_chain_html(sample_full_chain_matrices, tc_mapping=tc_mapping)

        assert "TC001" in html


# =============================================================================
# Test Records Markdown Tests
# =============================================================================


class TestRenderTestRecordsMarkdown:
    """Tests for test records Markdown renderer."""

    def test_renders_table(self, sample_test_records):
        """Test that Markdown output has table structure."""
        md = render_test_records_markdown(sample_test_records)

        assert "|" in md
        assert "---" in md

    def test_contains_test_id(self, sample_test_records):
        """Test that Markdown contains test ID."""
        md = render_test_records_markdown(sample_test_records)

        assert "TC001" in md

    def test_includes_metadata(self, sample_test_records, sample_metadata):
        """Test that Markdown includes metadata when provided."""
        md = render_test_records_markdown(sample_test_records, metadata=sample_metadata)

        assert "1.0.0" in md
        assert "CI Pipeline" in md


# =============================================================================
# Full Chain Markdown Tests
# =============================================================================


class TestRenderFullChainMarkdown:
    """Tests for full chain Markdown renderer."""

    def test_renders_table(self, sample_full_chain_matrices):
        """Test that Markdown output has table structure."""
        md = render_full_chain_markdown(sample_full_chain_matrices)

        assert "|" in md
        assert "---" in md

    def test_contains_document_columns(self, sample_full_chain_matrices):
        """Test that Markdown contains document hierarchy columns."""
        md = render_full_chain_markdown(sample_full_chain_matrices)

        assert "SYS" in md
        assert "SRS" in md


# =============================================================================
# Test Records JSON Tests
# =============================================================================


class TestRenderTestRecordsJson:
    """Tests for test records JSON renderer."""

    def test_valid_json(self, sample_test_records):
        """Test that output is valid JSON."""
        output = render_test_records_json(sample_test_records)

        data = json.loads(output)
        assert "summary" in data
        assert "tests" in data

    def test_contains_test_records(self, sample_test_records):
        """Test that JSON contains test records."""
        output = render_test_records_json(sample_test_records)

        data = json.loads(output)
        assert len(data["tests"]) == 2
        assert data["tests"][0]["test_id"] == "TC001"

    def test_summary_stats(self, sample_test_records):
        """Test that JSON contains summary statistics."""
        output = render_test_records_json(sample_test_records)

        data = json.loads(output)
        assert data["summary"]["total_tests"] == 2
        assert data["summary"]["passed"] == 1
        assert data["summary"]["failed"] == 1

    def test_includes_metadata(self, sample_test_records, sample_metadata):
        """Test that JSON includes metadata when provided."""
        output = render_test_records_json(sample_test_records, metadata=sample_metadata)

        data = json.loads(output)
        assert "metadata" in data
        assert data["metadata"]["software_version"] == "1.0.0"

    def test_metadata_with_none_environment(self, sample_test_records):
        """Test that JSON handles metadata with environment=None."""
        metadata = MatrixMetadata(
            software_version="1.0.0",
            tester_id="CI",
            execution_timestamp="2024-01-01T10:00:00Z",
            environment=None,  # No environment info
        )
        output = render_test_records_json(sample_test_records, metadata=metadata)

        data = json.loads(output)
        assert "metadata" in data
        assert data["metadata"]["software_version"] == "1.0.0"
        assert data["metadata"]["environment"] is None
        assert data["metadata"]["test_tools"] is None


# =============================================================================
# Full Chain JSON Tests
# =============================================================================


class TestRenderFullChainJson:
    """Tests for full chain JSON renderer."""

    def test_valid_json(self, sample_full_chain_matrices):
        """Test that output is valid JSON."""
        output = render_full_chain_json(sample_full_chain_matrices)

        data = json.loads(output)
        assert "summary" in data
        assert "matrices" in data

    def test_contains_matrices(self, sample_full_chain_matrices):
        """Test that JSON contains matrices."""
        output = render_full_chain_json(sample_full_chain_matrices)

        data = json.loads(output)
        assert len(data["matrices"]) == 1
        assert data["matrices"][0]["path_name"] == "SYS -> SRS"


# =============================================================================
# Test Records CSV Tests
# =============================================================================


class TestRenderTestRecordsCsv:
    """Tests for test records CSV renderer."""

    def test_valid_csv(self, sample_test_records):
        """Test that output is valid CSV."""
        output = render_test_records_csv(sample_test_records)

        reader = csv.reader(io.StringIO(output))
        rows = list(reader)
        assert len(rows) > 0

    def test_contains_test_records(self, sample_test_records):
        """Test that CSV contains test records."""
        output = render_test_records_csv(sample_test_records)

        assert "TC001" in output
        assert "test_valid" in output


# =============================================================================
# Full Chain CSV Tests
# =============================================================================


class TestRenderFullChainCsv:
    """Tests for full chain CSV renderer."""

    def test_valid_csv(self, sample_full_chain_matrices):
        """Test that output is valid CSV."""
        output = render_full_chain_csv(sample_full_chain_matrices)

        reader = csv.reader(io.StringIO(output))
        rows = list(reader)
        assert len(rows) > 0

    def test_contains_document_columns(self, sample_full_chain_matrices):
        """Test that CSV contains document hierarchy columns."""
        output = render_full_chain_csv(sample_full_chain_matrices)

        assert "SYS" in output
        assert "SRS" in output


# =============================================================================
# Test Records XLSX Tests
# =============================================================================


class TestRenderTestRecordsXlsx:
    """Tests for test records XLSX renderer."""

    def test_valid_xlsx(self, sample_test_records):
        """Test that output is valid XLSX."""
        output = render_test_records_xlsx(sample_test_records)

        wb = load_workbook(io.BytesIO(output))
        assert wb.active is not None

    def test_sheet_title(self, sample_test_records):
        """Test that sheet has correct title."""
        output = render_test_records_xlsx(sample_test_records)

        wb = load_workbook(io.BytesIO(output))
        assert wb.active.title == "Test Records"

    def test_includes_metadata_with_environment(self, sample_test_records, sample_metadata):
        """Test that metadata with environment is included."""
        output = render_test_records_xlsx(sample_test_records, metadata=sample_metadata)

        wb = load_workbook(io.BytesIO(output))
        ws = wb.active

        # Check metadata cells are present
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Software Version:" in values
        assert "1.0.0" in values
        assert "Environment:" in values
        assert "Test Tools:" in values

    def test_all_outcome_colors(self):
        """Test that all test outcomes get appropriate colors."""
        records = [
            TestRecord(
                test_id=f"TC{i:03d}",
                test_name=f"test_{outcome}",
                test_nodeid=f"test.py::test_{outcome}",
                outcome=outcome,
                requirements=["SRS001"],
            )
            for i, outcome in enumerate(["passed", "failed", "skipped", "error", "unknown"], 1)
        ]

        output = render_test_records_xlsx(records)
        wb = load_workbook(io.BytesIO(output))
        assert wb.active is not None


# =============================================================================
# Full Chain XLSX Tests
# =============================================================================


class TestRenderFullChainXlsx:
    """Tests for full chain XLSX renderer."""

    def test_valid_xlsx(self, sample_full_chain_matrices):
        """Test that output is valid XLSX."""
        output = render_full_chain_xlsx(sample_full_chain_matrices)

        wb = load_workbook(io.BytesIO(output))
        assert wb.active is not None

    def test_sheet_title(self, sample_full_chain_matrices):
        """Test that sheet has correct title."""
        output = render_full_chain_xlsx(sample_full_chain_matrices)

        wb = load_workbook(io.BytesIO(output))
        assert wb.active.title == "Trace Matrix"

    def test_multiple_matrices_create_sheets(self):
        """Test that multiple matrices create separate sheets."""
        item1 = Item(uid="UN001", text="User need", document_prefix="UN")
        item2 = Item(uid="SYS001", text="System req", document_prefix="SYS")

        matrices = [
            FullChainMatrix(
                path_name="UN -> SYS",
                document_hierarchy=["UN", "SYS"],
                rows=[ChainRow(chain={"UN": item1, "SYS": item2})],
                summary={"total": 1, "passed": 1},
            ),
            FullChainMatrix(
                path_name="UN -> HAZ",
                document_hierarchy=["UN", "HAZ"],
                rows=[ChainRow(chain={"UN": item1})],
                summary={"total": 1, "not_covered": 1},
            ),
        ]

        output = render_full_chain_xlsx(matrices)
        wb = load_workbook(io.BytesIO(output))

        # Should have 2 sheets
        assert len(wb.sheetnames) == 2

    def test_all_status_colors(self):
        """Test that all status values get appropriate colors."""
        item = Item(uid="SRS001", text="Req", document_prefix="SRS")

        matrices = [
            FullChainMatrix(
                path_name="SRS",
                document_hierarchy=["SRS"],
                rows=[
                    ChainRow(chain={"SRS": item}, rollup_status="Passed"),
                    ChainRow(chain={"SRS": item}, rollup_status="Failed"),
                    ChainRow(chain={"SRS": item}, rollup_status="Partial"),
                    ChainRow(chain={"SRS": item}, rollup_status="Not Covered"),
                    ChainRow(chain={"SRS": item}, rollup_status="N/A"),
                ],
                summary={"total": 5, "passed": 1, "failed": 1, "not_covered": 1},
            )
        ]

        output = render_full_chain_xlsx(matrices)
        wb = load_workbook(io.BytesIO(output))
        assert wb.active is not None

    def test_include_ancestors_column(self):
        """Test that ancestors column is included when flag is set."""
        item = Item(uid="SRS001", text="Req", document_prefix="SRS")

        matrices = [
            FullChainMatrix(
                path_name="SRS",
                document_hierarchy=["SRS"],
                rows=[ChainRow(chain={"SRS": item}, ancestor_uids=["UN001", "SYS001"])],
                summary={"total": 1},
                include_ancestors=True,
            )
        ]

        output = render_full_chain_xlsx(matrices)
        wb = load_workbook(io.BytesIO(output))
        ws = wb.active

        # Check for "Traces To" header
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Traces To" in values
