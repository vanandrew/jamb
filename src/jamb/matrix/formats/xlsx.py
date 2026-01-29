"""Excel (XLSX) traceability matrix output."""

import io
import warnings

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from jamb.core.models import FullChainMatrix, MatrixMetadata, TestRecord

# Threshold for warning about large datasets
LARGE_DATASET_WARNING_THRESHOLD = 5000

# =============================================================================
# Color and Style Constants
# =============================================================================

# Header styling
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Inline font for rich text (bold)
BOLD_INLINE = InlineFont(b=True)

# Coverage status fills
PASSED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAILED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
UNCOVERED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NA_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Test outcome fills
SKIPPED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
UNKNOWN_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Full chain matrix fills
PARTIAL_FILL = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")


def _make_item_rich_text(uid: str, header: str | None, text: str) -> CellRichText:
    """Create rich text with bold UID/header and regular text.

    Args:
        uid: The item UID.
        header: Optional header text.
        text: The item body text.

    Returns:
        CellRichText with bold UID/header portion and regular text.
    """
    if header:
        # Format: **UID: header** - text
        bold_part = f"{uid}: {header}"
        return CellRichText(
            TextBlock(BOLD_INLINE, bold_part),
            f" - {text}",
        )
    else:
        # Format: **UID:** text
        bold_part = f"{uid}:"
        return CellRichText(
            TextBlock(BOLD_INLINE, bold_part),
            f" {text}",
        )


def render_test_records_xlsx(
    records: list[TestRecord],
    metadata: MatrixMetadata | None = None,
) -> bytes:
    """Render test records as Excel test records matrix.

    Args:
        records: List of TestRecord objects to render.
        metadata: Optional matrix metadata for IEC 62304 5.7.5 compliance.

    Returns:
        Bytes containing an XLSX workbook with a styled header, metadata
        section, summary section, and color-coded outcome cells.
    """
    if len(records) > LARGE_DATASET_WARNING_THRESHOLD:
        warnings.warn(
            f"Large matrix ({len(records)} rows) may use significant memory. "
            "Consider using CSV format for large datasets.",
            stacklevel=2,
        )

    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "Test Records"

    # Calculate stats
    total = len(records)
    passed = sum(1 for r in records if r.outcome == "passed")
    failed = sum(1 for r in records if r.outcome == "failed")
    skipped = sum(1 for r in records if r.outcome == "skipped")
    error = sum(1 for r in records if r.outcome == "error")
    pass_rate = (100 * passed / total) if total else 0

    # Title
    ws["A1"] = "Test Records"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")

    # Metadata section
    current_row = 3
    if metadata:
        ws.cell(row=current_row, column=1, value="Software Version:")
        ws.cell(row=current_row, column=2, value=metadata.software_version or "Unknown")
        current_row += 1

        ws.cell(row=current_row, column=1, value="Tester:")
        ws.cell(row=current_row, column=2, value=metadata.tester_id)
        current_row += 1

        ws.cell(row=current_row, column=1, value="Date:")
        ws.cell(row=current_row, column=2, value=metadata.execution_timestamp or "Unknown")
        current_row += 1

        if metadata.environment:
            env = metadata.environment
            env_str = (
                f"{env.os_name} {env.os_version}, Python {env.python_version}, "
                f"{env.platform}, {env.processor}, {env.hostname}, "
                f"{env.cpu_count} cores"
            )
            ws.cell(row=current_row, column=1, value="Environment:")
            ws.cell(row=current_row, column=2, value=env_str)
            current_row += 1

            if env.test_tools:
                tools = [f"{name} {ver}" for name, ver in sorted(env.test_tools.items())]
                ws.cell(row=current_row, column=1, value="Test Tools:")
                ws.cell(row=current_row, column=2, value=", ".join(tools))
                current_row += 1

        current_row += 1  # Empty row separator

    # Summary section
    ws.cell(row=current_row, column=1, value="Total Tests:")
    ws.cell(row=current_row, column=2, value=total)
    current_row += 1
    ws.cell(row=current_row, column=1, value="Passed:")
    ws.cell(row=current_row, column=2, value=passed)
    current_row += 1
    ws.cell(row=current_row, column=1, value="Failed:")
    ws.cell(row=current_row, column=2, value=failed)
    current_row += 1
    ws.cell(row=current_row, column=1, value="Skipped:")
    ws.cell(row=current_row, column=2, value=skipped)
    current_row += 1
    ws.cell(row=current_row, column=1, value="Error:")
    ws.cell(row=current_row, column=2, value=error)
    current_row += 1
    ws.cell(row=current_row, column=1, value="Pass Rate:")
    ws.cell(row=current_row, column=2, value=f"{pass_rate:.1f}%")
    current_row += 2  # Extra row before header

    # Header row
    header_widths = [
        ("Test Case", 12),
        ("Test Name", 40),
        ("Outcome", 12),
        ("Requirements", 25),
        ("Test Actions", 40),
        ("Expected Results", 40),
        ("Actual Results", 40),
        ("Notes", 50),
        ("Timestamp", 22),
    ]
    headers = [h for h, _ in header_widths]
    header_row = current_row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    row = header_row + 1
    for rec in records:
        requirements_str = ", ".join(rec.requirements) if rec.requirements else ""
        test_actions_str = "\n".join(rec.test_actions) if rec.test_actions else ""
        expected_results_str = "\n".join(rec.expected_results) if rec.expected_results else ""
        actual_results_str = "\n".join(rec.actual_results) if rec.actual_results else ""
        notes_str = "\n".join(rec.notes) if rec.notes else ""

        # Determine fill color based on outcome
        outcome_lower = rec.outcome.lower() if rec.outcome else "unknown"
        if outcome_lower == "passed":
            fill = PASSED_FILL
        elif outcome_lower == "failed":
            fill = FAILED_FILL
        elif outcome_lower == "skipped":
            fill = SKIPPED_FILL
        elif outcome_lower == "error":
            fill = ERROR_FILL
        else:
            fill = UNKNOWN_FILL

        # Write row
        ws.cell(row=row, column=1, value=rec.test_id)
        ws.cell(row=row, column=2, value=rec.test_name)
        outcome_cell = ws.cell(row=row, column=3, value=rec.outcome)
        outcome_cell.fill = fill
        ws.cell(row=row, column=4, value=requirements_str)
        actions_cell = ws.cell(row=row, column=5, value=test_actions_str)
        actions_cell.alignment = Alignment(wrap_text=True, vertical="top")
        expected_cell = ws.cell(row=row, column=6, value=expected_results_str)
        expected_cell.alignment = Alignment(wrap_text=True, vertical="top")
        actual_cell = ws.cell(row=row, column=7, value=actual_results_str)
        actual_cell.alignment = Alignment(wrap_text=True, vertical="top")
        notes_cell = ws.cell(row=row, column=8, value=notes_str)
        notes_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=9, value=rec.execution_timestamp or "")

        row += 1

    # Auto-adjust column widths
    for col, (_, width) in enumerate(header_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def render_full_chain_xlsx(
    matrices: list[FullChainMatrix],
    tc_mapping: dict[str, str] | None = None,
) -> bytes:
    """Render full chain trace matrices as Excel workbook.

    Args:
        matrices: List of FullChainMatrix objects to render.
        tc_mapping: Optional mapping from test nodeid to TC ID for display.

    Returns:
        Bytes containing an XLSX workbook with all matrices.
    """
    total_rows = sum(len(m.rows) for m in matrices)
    if total_rows > LARGE_DATASET_WARNING_THRESHOLD:
        warnings.warn(
            f"Large matrix ({total_rows} rows) may use significant memory. "
            "Consider using CSV format for large datasets.",
            stacklevel=2,
        )

    tc_mapping = tc_mapping or {}
    wb = Workbook()

    # Overall summary
    total = sum(m.summary.get("total", 0) for m in matrices)
    passed = sum(m.summary.get("passed", 0) for m in matrices)
    failed = sum(m.summary.get("failed", 0) for m in matrices)
    not_covered = sum(m.summary.get("not_covered", 0) for m in matrices)

    # Create a sheet for each matrix path
    for i, matrix in enumerate(matrices):
        if i == 0:
            ws: Worksheet = wb.active  # type: ignore[assignment]
        else:
            ws = wb.create_sheet()

        # Sheet title (limited to 31 chars for Excel)
        # "Trace Matrix" = 12 chars, "Trace Matrix 999" = 16 chars, well under 31
        base_title = "Trace Matrix"
        ws.title = base_title if i == 0 else f"{base_title} {i + 1}"

        # Title
        ws["A1"] = "Traceability Matrix"
        ws["A1"].font = Font(bold=True, size=14)

        # Summary section (only on first sheet)
        current_row = 3
        if i == 0:
            ws.cell(row=current_row, column=1, value="Total Items:")
            ws.cell(row=current_row, column=2, value=total)
            current_row += 1
            ws.cell(row=current_row, column=1, value="Passed:")
            ws.cell(row=current_row, column=2, value=passed)
            current_row += 1
            ws.cell(row=current_row, column=1, value="Failed:")
            ws.cell(row=current_row, column=2, value=failed)
            current_row += 1
            ws.cell(row=current_row, column=1, value="Not Covered:")
            ws.cell(row=current_row, column=2, value=not_covered)
            current_row += 2

        # Build header row
        headers = []
        if matrix.include_ancestors:
            headers.append("Traces To")
        headers.extend(matrix.document_hierarchy)
        headers.extend(["Tests", "Status"])

        header_row = current_row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        row = header_row + 1
        for chain_row in matrix.rows:
            col = 1

            # Traces To column
            if matrix.include_ancestors:
                ws.cell(row=row, column=col, value=", ".join(chain_row.ancestor_uids))
                col += 1

            # Document columns
            for prefix in matrix.document_hierarchy:
                item = chain_row.chain.get(prefix)
                if item:
                    # Use rich text with bold UID/header
                    rich_text = _make_item_rich_text(item.uid, item.header, item.text)
                    ws.cell(row=row, column=col, value=rich_text)
                else:
                    ws.cell(row=row, column=col, value="")
                col += 1

            # Tests column
            tests = []
            for test in chain_row.descendant_tests:
                test_name = test.test_nodeid.split("::")[-1]
                outcome = test.test_outcome or "unknown"
                tc_id = tc_mapping.get(test.test_nodeid, "")
                tc_prefix = f"{tc_id}: " if tc_id else ""
                tests.append(f"{tc_prefix}{test_name} [{outcome}]")
            ws.cell(row=row, column=col, value="\n".join(tests))
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
            col += 1

            # Status column with color
            status = chain_row.rollup_status
            status_lower = status.lower().replace(" ", "_")
            if status_lower == "passed":
                fill = PASSED_FILL
            elif status_lower == "failed":
                fill = FAILED_FILL
            elif status_lower == "partial":
                fill = PARTIAL_FILL
            elif status_lower == "not_covered":
                fill = UNCOVERED_FILL
            else:
                fill = NA_FILL

            status_cell = ws.cell(row=row, column=col, value=status)
            status_cell.fill = fill

            row += 1

        # Set variable column widths based on content type
        col_widths = {
            "Traces To": 25,
            "Tests": 50,
            "Status": 12,
        }
        default_width = 35  # For document columns

        for col_idx, header in enumerate(headers, start=1):
            width = col_widths.get(header, default_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
